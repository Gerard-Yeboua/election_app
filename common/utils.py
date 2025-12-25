# apps/common/utils.py
from django.core.mail import send_mail
from django.conf import settings
from django.template.loader import render_to_string
from django.utils.html import strip_tags


def send_notification_email(user, subject, template_name, context):
    """Envoie un email de notification"""
    
    context['user'] = user
    context['site_name'] = 'CEI - Gestion PV'
    
    html_message = render_to_string(template_name, context)
    plain_message = strip_tags(html_message)
    
    send_mail(
        subject=subject,
        message=plain_message,
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[user.email],
        html_message=html_message,
        fail_silently=False,
    )


def format_phone_number(phone):
    """Formate un numéro de téléphone"""
    # Supprimer les espaces et caractères spéciaux
    phone = ''.join(filter(str.isdigit, phone))
    
    # Format: +225 XX XX XX XX XX
    if len(phone) == 10:
        return f"+225 {phone[0:2]} {phone[2:4]} {phone[4:6]} {phone[6:8]} {phone[8:10]}"
    
    return phone


def calculate_distance(lat1, lon1, lat2, lon2):
    """Calcule la distance entre deux points GPS (en mètres)"""
    from math import radians, cos, sin, asin, sqrt
    
    # Rayon de la Terre en mètres
    R = 6371000
    
    # Convertir en radians
    lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
    
    # Formule de Haversine
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * asin(sqrt(a))
    
    return R * c


def generate_reference_number(prefix, model, field_name='numero_reference'):
    """Génère un numéro de référence unique"""
    from django.utils import timezone
    import random
    import string
    
    timestamp = timezone.now().strftime('%Y%m%d%H%M%S')
    random_str = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
    
    reference = f"{prefix}-{timestamp}-{random_str}"
    
    # Vérifier l'unicité
    while model.objects.filter(**{field_name: reference}).exists():
        random_str = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
        reference = f"{prefix}-{timestamp}-{random_str}"
    
    return reference


def export_to_excel(queryset, filename, columns):
    """Exporte un queryset vers Excel"""
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Font, Alignment
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export"
    
    # En-têtes
    for col, (header, field) in enumerate(columns.items(), 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Données
    for row, obj in enumerate(queryset, 2):
        for col, (header, field) in enumerate(columns.items(), 1):
            # Gérer les champs imbriqués (ex: 'user__email')
            value = obj
            for attr in field.split('__'):
                value = getattr(value, attr, '')
            
            ws.cell(row=row, column=col).value = str(value)
    
    # Ajuster les largeurs
    for col in range(1, len(columns) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Préparer la réponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response