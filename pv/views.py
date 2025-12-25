from django.shortcuts import render

# Create your views here.
# apps/pv/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponseForbidden, JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum
from django.utils import timezone
from datetime import timedelta

from pv.models import ProcesVerbal, Candidat, ResultatCandidat, HistoriqueValidation
from pv.forms import ProcesVerbalForm, ResultatCandidatFormSet, ValidationForm
from pv.services.validation_service import validation_service
from accounts.models import CheckIn


@login_required
def pv_list(request):
    """Liste des PV (Admin)"""
    if request.user.role not in ['ADMIN', 'SUPER_ADMIN']:
        return HttpResponseForbidden()
    
    pv_qs = request.user.get_pv_accessibles().select_related(
        'bureau_vote', 'superviseur', 'validateur'
    )
    
    # Filtres
    statut = request.GET.get('statut')
    region = request.GET.get('region')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    search = request.GET.get('search')
    
    if statut:
        pv_qs = pv_qs.filter(statut=statut)
    
    if region:
        pv_qs = pv_qs.filter(
            bureau_vote__lieu_vote__sous_prefecture__commune__departement__region_id=region
        )
    
    if date_from:
        pv_qs = pv_qs.filter(date_soumission__gte=date_from)
    
    if date_to:
        pv_qs = pv_qs.filter(date_soumission__lte=date_to)
    
    if search:
        pv_qs = pv_qs.filter(
            Q(numero_reference__icontains=search) |
            Q(bureau_vote__code_bv__icontains=search) |
            Q(superviseur__email__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(pv_qs.order_by('-date_soumission'), 50)
    page_number = request.GET.get('page')
    pv_page = paginator.get_page(page_number)
    
    # Statistiques
    stats = {
        'total': pv_qs.count(),
        'valides': pv_qs.filter(statut='VALIDE').count(),
        'en_attente': pv_qs.filter(statut='EN_ATTENTE').count(),
        'rejetes': pv_qs.filter(statut='REJETE').count(),
    }
    
    context = {
        'pv_list': pv_page,
        'stats': stats,
        'statut_filter': statut,
        'search_query': search,
    }
    
    return render(request, 'pv/pv_list.html', context)


@login_required
def pv_detail(request, pv_id):
    """Détail d'un PV"""
    pv = get_object_or_404(
        ProcesVerbal.objects.select_related('bureau_vote', 'superviseur', 'validateur'),
        id=pv_id
    )
    
    # Vérifier les permissions
    if request.user.role == 'SUPERVISEUR':
        if pv.superviseur != request.user:
            return HttpResponseForbidden()
    elif not request.user.peut_acceder_bureau(pv.bureau_vote):
        return HttpResponseForbidden()
    
    # Résultats
    resultats = pv.resultats.select_related('candidat').order_by('candidat__numero_ordre')
    
    # Historique de validation
    historique = pv.historique_validations.select_related('validateur').order_by('-date_action')
    
    context = {
        'pv': pv,
        'resultats': resultats,
        'historique': historique,
        'can_validate': request.user.peut_valider_pv_bureau(pv.bureau_vote),
    }
    
    return render(request, 'pv/pv_detail.html', context)


@login_required
def submit_pv(request):
    """Soumettre un PV (Superviseur)"""
    if request.user.role != 'SUPERVISEUR':
        return HttpResponseForbidden()
    
    if not request.user.bureau_vote:
        messages.error(request, "Vous n'êtes affecté à aucun bureau de vote")
        return redirect('dashboard:index')
    
    # Vérifier le check-in actif
    checkin = CheckIn.objects.filter(
        superviseur=request.user,
        is_active=True
    ).first()
    
    if not checkin:
        messages.error(request, "Vous devez effectuer un check-in avant de soumettre un PV")
        return redirect('accounts:checkin_create')
    
    # Vérifier s'il existe déjà un PV validé pour ce bureau
    pv_existant = ProcesVerbal.objects.filter(
        bureau_vote=request.user.bureau_vote,
        statut='VALIDE'
    ).first()
    
    if pv_existant:
        messages.warning(request, "Un PV validé existe déjà pour ce bureau")
        return redirect('pv:detail', pv_id=pv_existant.id)
    
    if request.method == 'POST':
        form = ProcesVerbalForm(request.POST, request.FILES)
        
        if form.is_valid():
            pv = form.save(commit=False)
            pv.bureau_vote = request.user.bureau_vote
            pv.superviseur = request.user
            pv.checkin = checkin
            
            # Validation automatique des cohérences
            pv.save()
            pv.validate_coherence()
            
            messages.success(request, f"PV {pv.numero_reference} soumis avec succès!")
            return redirect('pv:add_results', pv_id=pv.id)
        else:
            messages.error(request, "Veuillez corriger les erreurs dans le formulaire")
    else:
        # Pré-remplir avec les données du bureau
        initial = {
            'nombre_inscrits': request.user.bureau_vote.nombre_inscrits
        }
        form = ProcesVerbalForm(initial=initial)
    
    context = {
        'form': form,
        'bureau': request.user.bureau_vote,
    }
    
    return render(request, 'pv/submit_pv.html', context)


@login_required
def add_results(request, pv_id):
    """Ajouter les résultats par candidat"""
    pv = get_object_or_404(ProcesVerbal, id=pv_id)
    
    if pv.superviseur != request.user:
        return HttpResponseForbidden()
    
    # Vérifier si les résultats n'existent pas déjà
    if pv.resultats.exists():
        messages.warning(request, "Les résultats ont déjà été enregistrés")
        return redirect('pv:detail', pv_id=pv.id)
    
    candidats = Candidat.objects.actifs().order_by('numero_ordre')
    
    if request.method == 'POST':
        # Validation de la somme
        total_voix = 0
        erreurs = []
        
        for candidat in candidats:
            voix_str = request.POST.get(f'voix_{candidat.id}', '0')
            try:
                voix = int(voix_str)
                if voix < 0:
                    erreurs.append(f"Nombre de voix invalide pour {candidat.nom_complet}")
                total_voix += voix
            except ValueError:
                erreurs.append(f"Nombre de voix invalide pour {candidat.nom_complet}")
        
        if erreurs:
            for erreur in erreurs:
                messages.error(request, erreur)
        elif total_voix != pv.suffrages_exprimes:
            messages.error(
                request,
                f"La somme des voix ({total_voix}) ne correspond pas aux suffrages exprimés ({pv.suffrages_exprimes})"
            )
        else:
            # Créer les résultats
            for candidat in candidats:
                voix = int(request.POST.get(f'voix_{candidat.id}', 0))
                ResultatCandidat.objects.create(
                    proces_verbal=pv,
                    candidat=candidat,
                    nombre_voix=voix
                )
            
            # Marquer comme calculé
            pv.resultats_calcules = True
            pv.save()
            
            messages.success(request, "Résultats enregistrés avec succès!")
            return redirect('pv:detail', pv_id=pv.id)
    
    context = {
        'pv': pv,
        'candidats': candidats,
    }
    
    return render(request, 'pv/add_results.html', context)


@login_required
def validation_queue(request):
    """File d'attente de validation (Admin)"""
    if request.user.role not in ['ADMIN', 'SUPER_ADMIN']:
        return HttpResponseForbidden()
    
    pv_en_attente = request.user.get_pv_accessibles().filter(
        statut='EN_ATTENTE'
    ).select_related('bureau_vote', 'superviseur').order_by('-date_soumission')
    
    # Statistiques
    stats = {
        'total': pv_en_attente.count(),
        'avec_incoherence': pv_en_attente.filter(has_incoherence=True).count(),
        'en_retard': pv_en_attente.filter(
            date_soumission__lt=timezone.now() - timedelta(hours=2)
        ).count(),
    }
    
    context = {
        'pv_list': pv_en_attente[:100],  # Limiter à 100
        'stats': stats,
    }
    
    return render(request, 'pv/validation_queue.html', context)


@login_required
def validate_pv(request, pv_id):
    """Valider un PV (Admin)"""
    if request.user.role not in ['ADMIN', 'SUPER_ADMIN']:
        return HttpResponseForbidden()
    
    pv = get_object_or_404(
        ProcesVerbal.objects.select_related('bureau_vote', 'superviseur'),
        id=pv_id
    )
    
    if not request.user.peut_valider_pv_bureau(pv.bureau_vote):
        return HttpResponseForbidden()
    
    if pv.statut != 'EN_ATTENTE':
        messages.warning(request, "Ce PV n'est pas en attente de validation")
        return redirect('pv:detail', pv_id=pv.id)
    
    if request.method == 'POST':
        form = ValidationForm(request.POST)
        
        if form.is_valid():
            action = form.cleaned_data['action']
            commentaires = form.cleaned_data.get('commentaires', '')
            motif_rejet = form.cleaned_data.get('motif_rejet', '')
            
            try:
                if action == 'valider':
                    validation_service.valider_pv(pv, request.user, commentaires)
                    messages.success(request, f"PV {pv.numero_reference} validé avec succès!")
                
                elif action == 'rejeter':
                    if not motif_rejet:
                        messages.error(request, "Le motif de rejet est obligatoire")
                        return render(request, 'pv/validate_pv.html', {
                            'pv': pv,
                            'form': form,
                            'resultats': pv.resultats.all()
                        })
                    
                    validation_service.rejeter_pv(pv, request.user, motif_rejet)
                    messages.success(request, f"PV {pv.numero_reference} rejeté")
                
                elif action == 'correction':
                    validation_service.demander_correction(pv, request.user, commentaires)
                    messages.success(request, "Correction demandée")
                
                return redirect('pv:validation_queue')
            
            except Exception as e:
                messages.error(request, f"Erreur: {str(e)}")
    else:
        form = ValidationForm()
    
    # Résultats
    resultats = pv.resultats.select_related('candidat').order_by('candidat__numero_ordre')
    
    context = {
        'pv': pv,
        'form': form,
        'resultats': resultats,
    }
    
    return render(request, 'pv/validate_pv.html', context)


@login_required
def my_pv_list(request):
    """Mes PV (Superviseur)"""
    if request.user.role != 'SUPERVISEUR':
        return HttpResponseForbidden()
    
    mes_pv = ProcesVerbal.objects.filter(
        superviseur=request.user
    ).select_related('bureau_vote').order_by('-date_soumission')
    
    # Statistiques
    stats = {
        'total': mes_pv.count(),
        'valides': mes_pv.filter(statut='VALIDE').count(),
        'en_attente': mes_pv.filter(statut='EN_ATTENTE').count(),
        'rejetes': mes_pv.filter(statut='REJETE').count(),
    }
    
    context = {
        'pv_list': mes_pv,
        'stats': stats,
    }
    
    return render(request, 'pv/my_pv_list.html', context)


@login_required
def candidat_list(request):
    """Liste des candidats"""
    candidats = Candidat.objects.all().order_by('numero_ordre')
    
    # Ajouter les résultats si admin
    if request.user.role in ['ADMIN', 'SUPER_ADMIN']:
        candidats = candidats.annotate(
            total_voix=Sum('resultats__nombre_voix', filter=Q(resultats__proces_verbal__statut='VALIDE')),
            nb_bureaux=Count('resultats__proces_verbal', filter=Q(resultats__proces_verbal__statut='VALIDE'), distinct=True)
        )
    
    context = {
        'candidats': candidats,
    }
    
    return render(request, 'pv/candidat_list.html', context)


@login_required
def candidat_detail(request, candidat_id):
    """Détail d'un candidat"""
    candidat = get_object_or_404(Candidat, id=candidat_id)
    
    # Résultats nationaux
    resultats_nationaux = candidat.get_resultats_nationaux()
    
    # Résultats par région
    resultats_regions = candidat.get_resultats_par_region()
    
    # Meilleurs scores
    meilleurs_scores = candidat.get_meilleurs_scores(limit=10)
    
    context = {
        'candidat': candidat,
        'resultats_nationaux': resultats_nationaux,
        'resultats_regions': resultats_regions,
        'meilleurs_scores': meilleurs_scores,
    }
    
    return render(request, 'pv/candidat_detail.html', context)


@login_required
def pv_delete(request, pv_id):
    """Supprimer un PV (uniquement si en attente et par le superviseur)"""
    pv = get_object_or_404(ProcesVerbal, id=pv_id)
    
    # Vérifications
    if pv.superviseur != request.user:
        return HttpResponseForbidden()
    
    if pv.statut != 'EN_ATTENTE':
        messages.error(request, "Seuls les PV en attente peuvent être supprimés")
        return redirect('pv:detail', pv_id=pv.id)
    
    if request.method == 'POST':
        numero_ref = pv.numero_reference
        pv.delete()
        messages.success(request, f"PV {numero_ref} supprimé avec succès")
        return redirect('pv:my_pv_list')
    
    return render(request, 'pv/pv_confirm_delete.html', {'pv': pv})


@login_required
def pv_export(request):
    """Exporter les PV en Excel"""
    if request.user.role not in ['ADMIN', 'SUPER_ADMIN']:
        return HttpResponseForbidden()
    
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Font, Alignment
    
    # Créer un workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PV Validés"
    
    # En-têtes
    headers = [
        'N° Référence', 'Bureau', 'Superviseur', 'Date Soumission',
        'Inscrits', 'Votants', 'Exprimés', 'Nuls', 'Blancs',
        'Taux Participation', 'Validateur', 'Date Validation'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Données
    pv_list = request.user.get_pv_accessibles().filter(statut='VALIDE').select_related(
        'bureau_vote', 'superviseur', 'validateur'
    )
    
    for row, pv in enumerate(pv_list, 2):
        ws.cell(row=row, column=1).value = pv.numero_reference
        ws.cell(row=row, column=2).value = pv.bureau_vote.code_bv
        ws.cell(row=row, column=3).value = pv.superviseur.nom_complet
        ws.cell(row=row, column=4).value = pv.date_soumission.strftime('%d/%m/%Y %H:%M')
        ws.cell(row=row, column=5).value = pv.nombre_inscrits
        ws.cell(row=row, column=6).value = pv.nombre_votants
        ws.cell(row=row, column=7).value = pv.suffrages_exprimes
        ws.cell(row=row, column=8).value = pv.bulletins_nuls
        ws.cell(row=row, column=9).value = pv.bulletins_blancs
        ws.cell(row=row, column=10).value = f"{pv.taux_participation}%"
        ws.cell(row=row, column=11).value = pv.validateur.nom_complet if pv.validateur else ''
        ws.cell(row=row, column=12).value = pv.date_validation.strftime('%d/%m/%Y %H:%M') if pv.date_validation else ''
    
    # Ajuster les largeurs
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Préparer la réponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=pv_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response